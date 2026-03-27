from __future__ import annotations

import openpyxl
from pathlib import Path
from typing import List, Dict, Any, TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd

# ── Noms de colonnes (correspondant aux en-têtes Excel) ──────────────
COL_NUM_PROJET   = "N° Projet"
COL_DATE         = "Date"
COL_NOM_PROJET   = "Nom projet"
COL_NUM_MACHINE  = "N° Machine"
COL_CLIENT       = "Client"
COL_CLIENT_FINAL = "Client final"
COL_DESIGNATION  = "Désignation"
COL_REFERENCE    = "Référence"
COL_NBR_MACHINES = "Nbr machines"
COL_MW           = "MW"
COL_KV           = "KV"
COL_COS_PHI      = "Cos(phi)"
COL_HZ           = "Hz"
COL_TR_MIN       = "TR/MIN"
COL_DAL          = "DAL"
COL_LFER         = "LFER"
COL_NB_POLES     = "NB POLES"
COL_NB_ENCOCHES  = "NB ENCOCHES"
COL_IC           = "IC"
COL_IM           = "IM"
COL_IP           = "IP"
COL_EEX          = "EEX"
COL_TYPE_PRODUIT = "Type produit"
COL_PRODUIT      = "Produit"
COL_TYPE_AFFAIRE = "Type affaire"
COL_DAS          = "DAS"
COL_SECTEUR      = "Secteur"

# ── Catégories de recherche ──────────────────────────────────────────
STRING_FIELDS   = [COL_NUM_PROJET, COL_NOM_PROJET, COL_CLIENT, COL_CLIENT_FINAL, COL_DESIGNATION]
NUMERIC_FIELDS  = [COL_NBR_MACHINES, COL_MW, COL_KV, COL_COS_PHI, COL_HZ,
                   COL_TR_MIN, COL_DAL, COL_LFER, COL_NB_ENCOCHES]
DROPDOWN_FIELDS = [COL_IM, COL_EEX, COL_TYPE_PRODUIT, COL_PRODUIT,
                   COL_TYPE_AFFAIRE, COL_DAS, COL_SECTEUR]
PREFILL_FIELDS  = [COL_TYPE_PRODUIT, COL_PRODUIT, COL_TYPE_AFFAIRE, COL_DAS, COL_SECTEUR]

# Colonnes dont les codes doivent être traduits en labels dans l'affichage
LABEL_MAP_COLUMNS = [COL_TYPE_PRODUIT, COL_PRODUIT, COL_TYPE_AFFAIRE, COL_DAS, COL_SECTEUR]

# Colonne(s) à ne pas afficher dans les résultats
HIDDEN_COLUMNS = ["Heures projet"]

# Colonnes de la feuille Projets (heures par code job)
PROJET_HOURS_COLUMNS = ["230ETELEC", "230ETMECA", "230ETMECNC", "230ETNQ",
                        "230ETREGU", "240RD", "240RDNC", "Total général"]


class MachineDatabase:
    """Charge et interroge la base de machines à partir d'un fichier Excel."""

    def __init__(self, filepath: str):
        import pandas as pd
        self.filepath = filepath
        self.df: pd.DataFrame = pd.DataFrame()
        self.df_projets: pd.DataFrame = pd.DataFrame()
        self.unique_values: Dict[str, List[str]] = {}
        self._loaded = False

    # ── Chargement ───────────────────────────────────────────────────
    def load(self) -> bool:
        import pandas as pd
        path = Path(self.filepath)
        if not path.exists():
            print(f"Fichier base machines non trouvé : {self.filepath}")
            return False
        try:
            self.df = pd.read_excel(self.filepath, sheet_name="Machines")
            self._load_projets_sheet()
            self._normalize_ip()
            self._extract_unique_values()
            self._loaded = True
            return True
        except Exception as e:
            print(f"Erreur lors du chargement de la base machines : {e}")
            return False

    @property
    def is_loaded(self) -> bool:
        return self._loaded and not self.df.empty

    def _load_projets_sheet(self):
        """Charge la feuille Projets (heures par code job par projet)."""
        import pandas as pd
        try:
            raw = pd.read_excel(self.filepath, sheet_name="Projets", header=None)
            # Column 0 = Projet ID, columns 1-7 = job codes, column 8 = Total
            cols = ["Projet"] + PROJET_HOURS_COLUMNS
            self.df_projets = raw.iloc[1:, :len(cols)].copy()
            self.df_projets.columns = cols
            self.df_projets["Projet"] = self.df_projets["Projet"].astype(str).str.strip()
            self.df_projets = self.df_projets.reset_index(drop=True)
        except Exception as e:
            print(f"Erreur lors du chargement de la feuille Projets : {e}")
            self.df_projets = pd.DataFrame()

    def _normalize_ip(self):
        """Convertit la colonne IP en chaînes propres ('23', '55', …)."""
        import pandas as pd
        if COL_IP not in self.df.columns:
            return
        def _to_ip_str(v):
            if pd.isna(v):
                return v          # garder NaN
            if isinstance(v, float):
                return str(int(v))
            return str(v).strip()
        self.df[COL_IP] = self.df[COL_IP].apply(_to_ip_str)

    def _extract_unique_values(self):
        """Prépare les listes de valeurs uniques pour les filtres."""
        import pandas as pd
        # Champs dropdown
        for col in DROPDOWN_FIELDS:
            if col in self.df.columns:
                vals = self.df[col].dropna().astype(str).str.strip()
                vals = sorted(vals[vals != ""].unique())
                self.unique_values[col] = vals

        # Années
        if COL_DATE in self.df.columns:
            dates = pd.to_datetime(self.df[COL_DATE], errors="coerce")
            years = dates.dt.year.dropna().astype(int).unique().tolist()
            years.sort(reverse=True)
            self.unique_values[COL_DATE] = [str(y) for y in years]

        # Chiffres IP
        if COL_IP in self.df.columns:
            ips = self.df[COL_IP].dropna().astype(str)
            ips = ips[ips.str.len() >= 2]
            self.unique_values["IP_first"]  = sorted({ip[0] for ip in ips if ip[0].isdigit()})
            self.unique_values["IP_second"] = sorted({ip[1] for ip in ips if ip[1].isdigit()})

    # ── Recherche ────────────────────────────────────────────────────
    def search(self, filters: Dict[str, Any], tolerance_percent: float = 10.0) -> pd.DataFrame:
        """Filtre la base selon *filters*.

        Règle : si une cellule de la base est vide/NaN pour un champ filtré,
        la ligne n'est **pas** exclue (les cases vides sont toujours incluses).
        """
        import pandas as pd
        if self.df.empty:
            return self.df.copy()

        mask = pd.Series(True, index=self.df.index)

        for field, value in filters.items():
            if value is None or (isinstance(value, str) and value.strip() in ("", "Tous")):
                continue

            # Déterminer la colonne réelle
            col = COL_IP if field in ("IP_first", "IP_second") else field
            if col not in self.df.columns:
                continue

            if field in STRING_FIELDS:
                is_empty = self.df[col].isna() | (self.df[col].astype(str).str.strip() == "")
                matches  = self.df[col].astype(str).str.lower().str.contains(
                    str(value).lower(), na=False, regex=False
                )
                mask &= is_empty | matches

            elif field == COL_DATE:
                dates    = pd.to_datetime(self.df[col], errors="coerce")
                is_empty = dates.isna()
                try:
                    matches = dates.dt.year == int(value)
                except (ValueError, TypeError):
                    continue
                mask &= is_empty | matches

            elif field in NUMERIC_FIELDS:
                try:
                    target = float(value)
                except (ValueError, TypeError):
                    continue
                col_num  = pd.to_numeric(self.df[col], errors="coerce")
                is_empty = col_num.isna()
                tol      = abs(target) * tolerance_percent / 100.0
                matches  = (col_num >= target - tol) & (col_num <= target + tol)
                mask &= is_empty | matches

            elif field == COL_NB_POLES:
                col_num  = pd.to_numeric(self.df[col], errors="coerce")
                is_empty = col_num.isna()
                if value == ">4":
                    matches = col_num > 4
                else:
                    try:
                        matches = col_num == int(value)
                    except (ValueError, TypeError):
                        continue
                mask &= is_empty | matches

            elif field == "IP_first":
                if value == "x":
                    continue
                ip_str   = self.df[col].astype(str)
                is_empty = self.df[col].isna() | (ip_str.isin(["", "nan"]))
                matches  = ip_str.str[0] == str(value)
                mask &= is_empty | matches

            elif field == "IP_second":
                if value == "x":
                    continue
                ip_str   = self.df[col].astype(str)
                is_empty = self.df[col].isna() | (ip_str.isin(["", "nan"]))
                matches  = ip_str.str[1] == str(value)
                mask &= is_empty | matches

            elif field in DROPDOWN_FIELDS:
                col_vals = self.df[col].astype(str).str.strip()
                is_empty = self.df[col].isna() | col_vals.isin(["", "nan"])
                matches  = col_vals == str(value)
                mask &= is_empty | matches

        return self.df[mask].reset_index(drop=True)

    # ── Données projet (double-clic) ─────────────────────────────────
    def get_project_machines(self, project_id: str) -> pd.DataFrame:
        """Retourne toutes les machines appartenant au même projet."""
        import pandas as pd
        if self.df.empty or COL_NUM_PROJET not in self.df.columns:
            return pd.DataFrame()
        mask = self.df[COL_NUM_PROJET].astype(str).str.strip() == project_id.strip()
        return self.df[mask].reset_index(drop=True)

    def get_project_hours(self, project_id: str) -> Dict[str, Any]:
        """Retourne les heures du projet (ventilation par code job)."""
        if self.df_projets.empty:
            return {}
        row = self.df_projets[self.df_projets["Projet"] == project_id.strip()]
        if row.empty:
            return {}
        import pandas as pd
        result = {}
        for col in PROJET_HOURS_COLUMNS:
            val = row.iloc[0][col]
            result[col] = val if pd.notna(val) else 0.0
        return result

    # ── Modification d'une cellule ───────────────────────────────────
    def update_machine_cell(self, df_index: int, column: str, value) -> bool:
        """Met à jour une cellule dans le DataFrame en mémoire ET dans le fichier Excel.
        
        df_index : index dans self.df (pas l'index du sous-DataFrame filtré).
        """
        if column not in self.df.columns:
            return False
        # Gérer l'incompatibilité de type (ex: string dans colonne float64)
        import numpy as np
        if self.df[column].dtype.kind in ("f", "i", "u"):
            if value == "":
                value = np.nan
            elif isinstance(value, str):
                self.df[column] = self.df[column].astype(object)
        # Mise à jour en mémoire
        self.df.at[df_index, column] = value
        # Mise à jour dans le fichier Excel
        try:
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb["Machines"]
            # Trouver l'index de la colonne dans le fichier (1-based, row 1 = header)
            header_row = [cell.value for cell in ws[1]]
            # La colonne Excel peut avoir un nom légèrement différent — correspondance exacte
            excel_col_map = {
                COL_NUM_PROJET: "N° Projet", COL_DATE: "DATE",
                "Nom projet": "PROJET", COL_NUM_MACHINE: "NUMERO",
                COL_CLIENT: "CLIENT", COL_CLIENT_FINAL: "CLIENT FINAL",
                COL_DESIGNATION: "DESCRIPTION", COL_REFERENCE: "REFERENCE",
                COL_NBR_MACHINES: "NB MACHINES", COL_MW: "MW", COL_KV: "KV",
                COL_COS_PHI: "CPHI", COL_HZ: "HZ", COL_TR_MIN: "TR/MIN",
                COL_DAL: "DAL", COL_LFER: "LFER", COL_NB_POLES: "NB POLES",
                COL_NB_ENCOCHES: "NB ENCOCHES", COL_IC: "IC", COL_IM: "IM",
                COL_IP: "IP", COL_EEX: "EEX", COL_TYPE_PRODUIT: "Type produit",
                COL_PRODUIT: "Produit", COL_TYPE_AFFAIRE: "Type affaire",
                COL_DAS: "DAS", COL_SECTEUR: "Secteur",
            }
            excel_col_name = excel_col_map.get(column, column)
            if excel_col_name not in header_row:
                wb.close()
                return False
            col_idx = header_row.index(excel_col_name) + 1  # 1-based
            row_idx = df_index + 2  # 1-based, +1 header +1 for 0-based
            excel_value = None if (isinstance(value, float) and np.isnan(value)) else value
            ws.cell(row=row_idx, column=col_idx, value=excel_value)
            wb.save(self.filepath)
            wb.close()
            return True
        except Exception as e:
            print(f"Erreur sauvegarde Excel : {e}")
            return False

    def get_original_df_indices(self, project_id: str) -> list:
        """Retourne les indices du DataFrame principal pour un projet donné."""
        if self.df.empty or COL_NUM_PROJET not in self.df.columns:
            return []
        mask = self.df[COL_NUM_PROJET].astype(str).str.strip() == project_id.strip()
        return self.df[mask].index.tolist()
