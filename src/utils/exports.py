from __future__ import annotations
from copy import copy
import os
from typing import TYPE_CHECKING, Dict, List, Any

import openpyxl
from openpyxl.styles import Font

from src.utils.Task import GeneralTask

if TYPE_CHECKING:
    from src.model import Model, Project


# ── Helpers bas-niveau pour la feuille Excel ────────────────────────

def _auto_hours(task, ctx: dict) -> float:
    """Heures de base sans aucune correction manuelle (ni manuelle ni par catégorie)."""
    saved_manual = task.manual_base_hours
    saved_override = task.category_override_hours
    task.manual_base_hours = None
    task.category_override_hours = None
    h = task.effective_hours(ctx)
    task.manual_base_hours = saved_manual
    task.category_override_hours = saved_override
    return h


def _merge_col_b(ws, start: int, end: int, label: str):
    cell = ws.cell(row=start, column=2)
    cell.value = label
    cell.font = Font(bold=True)
    if end > start:
        ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)


def _write_d_e_f(ws, row: int, auto: float, effective: float, rex: float):
    """Écrit D = heures de base, E = heures finales (gras si corrigées), F = E × REX."""
    ws.cell(row=row, column=4).value = auto
    cell_e = ws.cell(row=row, column=5)
    cell_e.value = effective
    if effective != auto:
        cell_e.font = Font(bold=True)
    ws.cell(row=row, column=6).value = effective * rex


# ── En-tête projet ──────────────────────────────────────────────────

def _header_fields(project: "Project"):
    secteur_labels = {code: label for das, sectors in project.app_data.secteurs.items() for code, label in sectors.items()}
    product_labels = {code: label for mt, products in project.app_data.product.items() for code, label in products.items()}
    # row, attr, label_map
    return [
        (3,  "crm_number", None),
        (4,  "client", None),
        (5,  "affaire", project.app_data.types_affaires),
        (6,  "das", project.app_data.das),
        (7,  "secteur", secteur_labels),
        (8,  "machine_type", project.app_data.product_types),
        (9,  "product", product_labels),
        (10, "designation", None),
        (11, "quantity", None),
        (12, "revision", None),
        (13, "date", None),
        (14, "created_by", None),
        (15, "validated_by", None),
    ]


def _write_header(ws, project: "Project"):
    for row, attr, label_map in _header_fields(project):
        if label_map:
            value = getattr(project, attr)
            ws[f"C{row}"] = label_map.get(value, value)
        else:
            ws[f"C{row}"] = getattr(project, attr)
    ws["D4"] = project.description


# ── Points d'entrée publics ─────────────────────────────────────────

def export_ortems_excel(project: "Project", path: str):
    repartition = project.make_ortems_repartition()

    template_path = project.app_data.ortems_template_path
    wb = openpyxl.load_workbook(template_path)
    ws_ortems = wb["prepa ORTEMS"]
    col = 3
    job_labels = project.app_data.jobs
    for job, hours in repartition.items():
        ws_ortems.cell(row=1, column=col).value = job_labels.get(job, job)
        ws_ortems.cell(row=2, column=col).value = hours
        col += 1

    delai = project.compute_delai_etude()
    ws_ortems["B2"] = delai["delai_reel"]

    wb.active = ws_ortems
    wb.save(path)


def export_excel_report(project: "Project", path: str):
    template_path = project.app_data.excel_report_template_path
    wb = openpyxl.load_workbook(template_path)
    ws = wb['chiffrage']
    ctx = project.context()
    rex = project.manual_rex_coeff

    # Recalcul des totaux (cascade : nrc_subtotal, rc_subtotal, nrc_total, rc_total)
    project.compute_n_machines_total()
    project.calculate_total_with_rex()

    _write_header(ws, project)

    summary_tree = project.generate_summary_tree()

    # Enclenchement (rows 17-20)
    for r, task in zip(range(17, 21), summary_tree['Enclenchement']):
        _write_d_e_f(ws, r, task.default_hours(ctx), task.effective_hours(ctx), rex)

    # Calculs (rows 21-25, 1 ligne par catégorie)
    for r, (cat_label, calcul_list) in zip(range(21, 26), summary_tree['Calculs'].items()):
        active = [c for c in calcul_list if c.is_active(ctx)]
        ws.cell(row=r, column=3).value = cat_label
        _write_d_e_f(ws, r,
                     sum(c.default_hours(ctx) for c in active),
                     sum(c.effective_hours(ctx) for c in active), rex)

    # Plans fab (dynamique)
    row = 26
    fonts = [copy(ws.cell(row=21, column=c).font) for c in range(2, 7)]
    borders = [copy(ws.cell(row=21, column=c).border) for c in range(2, 7)]
    alignments = [copy(ws.cell(row=21, column=c).alignment) for c in range(2, 7)]
    number_formats = [copy(ws.cell(row=21, column=c).number_format) for c in range(2, 7)]
    fills = [copy(ws.cell(row=row, column=c).fill) for c in range(2, 7)]
    ws.delete_rows(row)

    tasks: Dict[str, List[GeneralTask]] = summary_tree['Plans / Specs / LDN']
    for subcat, tlist in tasks.items():
        if sum(t.effective_hours(ctx) for t in tlist) == 0:
            continue
        first_row = row
        for t in tlist:
            ws.insert_rows(row)
            ws.cell(row=row, column=3).value = t.label
            _write_d_e_f(ws, row, t.default_hours(ctx), t.effective_hours(ctx), rex)
            for c in range(2, 7):
                ws.cell(row=row, column=c).font = fonts[c-2]
                ws.cell(row=row, column=c).border = borders[c-2]
                ws.cell(row=row, column=c).fill = fills[c-2]
                ws.cell(row=row, column=c).alignment = alignments[c-2]
                ws.cell(row=row, column=c).number_format = number_formats[c-2]
            row += 1
        _merge_col_b(ws, first_row, row - 1, f"Plans FAB: {subcat}")

    # Options (dynamique)
    fills = [copy(ws.cell(row=row, column=c).fill) for c in range(2, 7)]
    ws.delete_rows(row)
    options: Dict[str, List[GeneralTask]] = summary_tree['Options']
    first_row = row
    for cat_label, option_list in options.items():
        if sum(o.effective_hours(ctx) for o in option_list) == 0:
            continue
        ws.insert_rows(row)
        ws.cell(row=row, column=3).value = cat_label
        _write_d_e_f(ws, row,
                     sum(o.default_hours(ctx) for o in option_list if o.is_active(ctx)),
                     sum(o.effective_hours(ctx) for o in option_list), rex)
        for c in range(2, 7):
            ws.cell(row=row, column=c).font = fonts[c-2]
            ws.cell(row=row, column=c).border = borders[c-2]
            ws.cell(row=row, column=c).fill = fills[c-2]
            ws.cell(row=row, column=c).alignment = alignments[c-2]
            ws.cell(row=row, column=c).number_format = number_formats[c-2]
        row += 1
    if row > first_row:
        _merge_col_b(ws, first_row, row - 1, "Options")

    # LPDC (2 lignes : BASE et PART)
    for cat_label, lpdc_list in summary_tree['Plans et documents contractuels'].items():
        active = [d for d in lpdc_list if d.is_active(ctx)]
        _write_d_e_f(ws, row,
                     sum(d.default_hours(ctx) for d in active),
                     sum(d.effective_hours(ctx) for d in active), rex)
        row += 1

    # LABO (2 lignes : LAB_METAL et LAB_ISOL)
    for cat_label, labo_list in summary_tree['Laboratoire'].items():
        active = [la for la in labo_list if la.is_active(ctx)]
        _write_d_e_f(ws, row,
                     sum(la.default_hours(ctx) for la in active),
                     sum(la.effective_hours(ctx) for la in active), rex)
        row += 1

    # NRC — Divers risques techniques
    nrc_divers = project.nrc_subtotal * project.divers_percent
    _write_d_e_f(ws, row, nrc_divers, nrc_divers, rex)
    row += 1

    # TOTAL NRC
    ws.cell(row=row, column=5).value = project.nrc_total
    ws.cell(row=row, column=6).value = project.nrc_total * rex
    row += 1

    # Suivi (6 lignes RC, labels déjà dans le template)
    for task in summary_tree['Suivi']:
        _write_d_e_f(ws, row, task.default_hours(ctx), task.effective_hours(ctx), rex)
        row += 1

    # RC — Divers risques techniques (pour 1 machine)
    rc_divers = project.rc_subtotal * project.divers_percent
    _write_d_e_f(ws, row, rc_divers, rc_divers, rex)
    row += 1

    # TOTAL RC (1 machine avec divers)
    rc_total_1machine = project.rc_subtotal * (1 + project.divers_percent)
    ws.cell(row=row, column=5).value = rc_total_1machine
    ws.cell(row=row, column=6).value = rc_total_1machine * rex
    row += 1

    # RC POUR N machines (heures RC totales après application du coefficient multi-machine)
    ws.cell(row=row, column=5).value = project.rc_total
    ws.cell(row=row, column=6).value = project.rc_total * rex
    row += 1

    # TOTAL RC + NRC
    ws.cell(row=row, column=5).value = project.n_machines_total
    ws.cell(row=row, column=6).value = project.total_with_rex

    wb.save(path)


def quick_export(model: "Model") -> Dict[str, str]:
    prj = model.project
    file_name = f"{prj.crm_number}{prj.revision}"

    # Garantit un dossier d'export valide, même si la config pointe sur le dossier parent.
    base_export_dir = prj.app_data.quick_export_path or prj.app_data.project_save_dir or "."
    base_export_dir = os.path.normpath(base_export_dir)
    if os.path.basename(base_export_dir).lower() != "chiffrages het":
        base_export_dir = os.path.join(base_export_dir, "Chiffrages HET")
    os.makedirs(base_export_dir, exist_ok=True)

    rapport_path = os.path.join(base_export_dir, f"{file_name}_rapport.xlsx")
    ortems_path = os.path.join(base_export_dir, f"{file_name}_ortems.xlsx")
    export_excel_report(prj, rapport_path)
    export_ortems_excel(prj, ortems_path)
    return {"rapport": rapport_path, "ortems": ortems_path}