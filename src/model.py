import copy
import pandas as pd
from typing import Dict, List, Optional, Any
from src.utils.ApplicationData import ApplicationData
from src.utils.Task import AbstractTask, GeneralTask, LPDCDocument, Labo, Option, Calcul
from PyQt6.QtCore import QObject, pyqtSignal
import openpyxl

class Project:
    def __init__(self, app_data: ApplicationData):
        self.app_data = app_data
        # Données générales
        self.crm_number = ""
        self.client = ""
        self.affaire = ""
        self.das = ""
        self.secteur = ""
        self.machine_type = ""
        self.product = ""
        self.designation = ""
        self.quantity = 1
        self.revision = "A"
        self.date = ""
        self.created_by = ""
        self.validated_by = ""
        self.description = ""

        self.lpdc_secteur_coeff: float = 1.0
        self.lpdc_affaire_coeff: float = 1.0
        self.calcul_coeff_type_affaire: Dict[str, float] = {} # Dict[activité: coeff] appliqué aux calculs selon le type d'affaire
        self.option_coeff_category: Dict[str, float] = {} # Dict[category: coeff] appliqué aux options selon le type d'affaire
        
        self.divers_percent: float = 0.05
        self.manual_rex_coeff: float = 1.0
        self.manual_rex_hours: Optional[float] = None

        self.first_machine_subtotal: Optional[float] = None
        self.first_machine_total: Optional[float] = None
        self.n_machines_total: Optional[float] = None
        self.total_with_rex: Optional[float] = None
        
        self.tasks: Dict[str, Dict[str, List[GeneralTask]]] = {}
        self.lpdc_docs: List[LPDCDocument] = []
        self.options: List[Option] = []
        self.calculs: List[Calcul] = []
        self.labo: List[Labo] = []
    
    def context(self) -> Dict[str, str]:
        return {
            "product": self.product,
            "machine_type": self.machine_type,
            "affaire": self.affaire,
            "secteur": self.secteur,

            "LPDC_secteur_coeff": self.lpdc_secteur_coeff,
            "LPDC_affaire_coeff": self.lpdc_affaire_coeff,
            "calcul_coeff_type_affaire": self.calcul_coeff_type_affaire,
            "option_coeff_category": self.option_coeff_category
        }

    def apply_defaults(self):
        """Applique les valeurs par défaut après avoir choisi le type de machine, le secteur et le type d'affaire."""
        
        ctx = self.context()

        # Récupérer les coefficients dépendant du contexte depuis app_data
        self.lpdc_secteur_coeff = self.app_data.lpdc_coeff_secteur.get(self.secteur, 1.0)
        self.lpdc_affaire_coeff = self.app_data.lpdc_coeff_affaire.get(self.affaire, 1.0)
        self.calcul_coeff_type_affaire = self.app_data.calcul_coeff_type_affaire.get(self.affaire, {})
        self.option_coeff_category = self.app_data.option_category_coeff.get(self.affaire, {})
        
        self.divers_percent = 0.05
        self.manual_rex_coeff = 1.0
        self.manual_rex_hours = None

        self.first_machine_subtotal = None
        self.first_machine_total = None
        self.n_machines_total = None
        self.total_with_rex = None
        
        self.tasks = copy.deepcopy(self.app_data.tasks)
        self.lpdc_docs = copy.deepcopy([doc for doc in self.app_data.lpdc_docs if doc.is_active(ctx) or doc.option_possible])
        self.options = copy.deepcopy(self.app_data.options)
        self.calculs = copy.deepcopy([calc for calc in self.app_data.calculs if calc.is_available_as_option(ctx) or calc.is_mandatory(ctx)])
        self.labo = copy.deepcopy(self.app_data.labo)
    
    def get_task_default_hours(self, task: GeneralTask) -> float:
        return task.default_hours(self.context())

    def get_all_tasks(self) -> List[GeneralTask]:
        """Retourne la liste plate de toutes les tâches générales."""
        return [task for subcats in self.tasks.values() for tasks in subcats.values() for task in tasks]
    
    def grouped_claculs(self) -> Dict[str, List[Calcul]]:
        grouped_calculs = {}
        for calc in self.calculs:
            if calc.category not in grouped_calculs:
                grouped_calculs[calc.category] = []
            grouped_calculs[calc.category].append(calc)
        return grouped_calculs
    
    def grouped_options(self) -> Dict[str, List[Option]]:
        grouped_options = {}
        for opt in self.options:
            if opt.category not in grouped_options:
                grouped_options[opt.category] = []
            grouped_options[opt.category].append(opt)
        return grouped_options
    
    def grouped_lpdc(self) -> Dict[str, List[LPDCDocument]]:
        """Retourne les documents LPDC regroupés en 'Base' et 'Particulières'."""
        grouped_lpdc = {"BASE": [], "PART": []} # ! Utiliser les codes catégories ("BASE", "PART")

        for doc in self.lpdc_docs:
            if self.machine_type not in doc.applicable_pour:
                continue
            if self.secteur in doc.secteur_obligatoire:
                grouped_lpdc["BASE"].append(doc)
            elif doc.option_possible:
                grouped_lpdc["PART"].append(doc)
        return grouped_lpdc
    
    def grouped_labo(self) -> Dict[str, List[Labo]]:
        grouped_labo = {}
        for labo in self.labo:
            if labo.category not in grouped_labo:
                grouped_labo[labo.category] = []
            grouped_labo[labo.category].append(labo)
        return grouped_labo

    def _change_group_label(self, grouped: Dict[str, List], label_map: Dict[str, str]) -> Dict[str, List]:
        """Change les labels des groupes selon un mapping fourni."""
        new_grouped = {}
        for old_label, items in grouped.items():
            new_label = label_map.get(old_label, old_label)
            new_grouped[new_label] = items
        return new_grouped

    def generate_summary_tree(self) -> Dict[str, Any]:
        return {
            "Tâches Générales": self.tasks,
            "Calculs": self._change_group_label(self.grouped_claculs(), self.app_data.calcul_categories),
            "Options": self._change_group_label(self.grouped_options(), self.app_data.option_categories),
            "Plans et documents contractuels": self._change_group_label(self.grouped_lpdc(), self.app_data.lpdc_categories),
            "Laboratoire": self._change_group_label(self.grouped_labo(), self.app_data.labo_categories),
        }
    
    def compute_tree_hours(self, node) -> float:
        if isinstance(node, AbstractTask):
            return node.effective_hours(self.context())
        if isinstance(node, list):
            return sum(self.compute_tree_hours(t) for t in node)
        if isinstance(node, dict):
            return sum(self.compute_tree_hours(v) for v in node.values())
        return 0.0

    def compute_first_machine_subtotal(self) -> float:
        """Calcule le sous-total de base (toutes les tâches)."""
        self.first_machine_subtotal = sum(
            self.compute_tree_hours(data)
            for data in [self.tasks, self.lpdc_docs, self.options, self.calculs, self.labo]
        )
        return self.first_machine_subtotal
    
    def compute_first_machine_total(self) -> float:
        """Calcule le total avec divers."""
        self.first_machine_total = self.first_machine_subtotal * (1 + self.divers_percent)
        return self.first_machine_total
    
    def _compute_multi_machine_coeff(self, quantity: int) -> float:
        """Calcule le coefficient pour machines multiples."""
        if quantity == 2:
            return 1.0
        elif quantity <= 5:
            return (quantity - 1) * 0.75
        elif quantity <= 25:
            return (quantity - 1) * 0.35
        else:
            return (quantity - 1) * 0.15
        
    def compute_n_machines_total(self) -> float:
        """Calcule le total pour n machines."""
        multiplicative_tasks_hours = sum([t.effective_hours(self.context()) for t in self.get_all_tasks() if t.mutiplicative])
        coeff = self._compute_multi_machine_coeff(self.quantity)
        additional_hours = multiplicative_tasks_hours * coeff

        self.n_machines_total = self.first_machine_total + additional_hours 
        return self.n_machines_total
    
    def calculate_total_with_rex(self) -> float:
        if self.manual_rex_hours is not None:
            self.total_with_rex = self.manual_rex_hours
        else:
            self.total_with_rex = self.n_machines_total * self.manual_rex_coeff
        return self.total_with_rex
    
    def _apply_group_repartition(
        self,
        repartition: Dict[str, float],
        grouped: Dict[str, List],
        ortems_map: Dict[str, Dict[str, float]],
    ) -> None:
        """Accumule les heures d'un groupe catégorisé dans repartition."""
        for category, items in grouped.items():
            rep = ortems_map.get(category, {})
            cat_sum = sum(item.effective_hours(self.context()) for item in items)
            for job_code, coeff in rep.items():
                repartition[job_code] += cat_sum * coeff

    def make_ortems_repartition(self) -> Dict[str, float]:
        """Crée la répartition ortems finale en combinant les différentes sources."""
        repartition = dict.fromkeys(self.app_data.jobs.keys(), 0.0)

        # Tâches générales (répartition définie par tâche)
        for task in self.get_all_tasks():
            hours = task.effective_hours(self.context())
            if hours > 0 and task.ortems_repartition:
                for job_code, coeff in task.ortems_repartition.items():
                    repartition[job_code] += hours * coeff

        # Sources catégorisées
        for grouped, ortems_map in [
            (self.grouped_claculs(), self.app_data.calcul_ortems),
            (self.grouped_options(), self.app_data.option_ortems),
            (self.grouped_lpdc(),    self.app_data.lpdc_ortems),
            (self.grouped_labo(),    self.app_data.labo_ortems),
        ]:
            self._apply_group_repartition(repartition, grouped, ortems_map)

        # Application du divers et du REX
        for job_code in repartition:
            repartition[job_code] *= (1 + self.divers_percent) * self.manual_rex_coeff


        return repartition
    
    def export_ortems_excel(self, path: str):
        """Exporte la répartition ortems dans un fichier Excel basé sur le template."""
        repartition = self.make_ortems_repartition()

        template_path = self.app_data.ortems_template_path
        wb = openpyxl.load_workbook(template_path)

        # --- Feuille 'prepa ORTEMS' : clés en ligne 1 (B:AX), valeurs en ligne 2 ---
        print("Répartition ORTEMS à exporter :")
        ws_ortems = wb["prepa ORTEMS"]
        col = 3
        job_labels = self.app_data.jobs
        for job, hours in repartition.items():
            print(f"    {job} : {hours:.1f} h")
            ws_ortems.cell(row=1, column=col).value = job_labels.get(job, job)
            ws_ortems.cell(row=2, column=col).value = hours
            col += 1

        # --- Feuille 'Calculs' : B3 = 1.2, C3 = 1.5 ---
        ws_calculs = wb["Calculs"]
        ws_calculs["B3"] = 1.2
        ws_calculs["C3"] = 1.5

        wb.save(path)


class Model(QObject):
    project_changed = pyqtSignal()  # Émis lors de l'application des paramètres par défaut
    data_updated = pyqtSignal()     # Émis lors de modifications mineures (valeurs, checkboxes)

    def __init__(self, app_data: ApplicationData):
        super().__init__()
        self.app_data = app_data
        self.project = Project(app_data)

    # ------------------------------------------------------------------
    # Sauvegarde / Chargement
    # ------------------------------------------------------------------

    def save_project(self) -> dict:
        """Sérialise le projet : valeurs scalaires + delta des modifications."""
        prj = self.project
        return {
            "version": 1,
            "project": {
                "crm_number":   prj.crm_number,
                "client":       prj.client,
                "affaire":      prj.affaire,
                "das":          prj.das,
                "secteur":      prj.secteur,
                "machine_type": prj.machine_type,
                "product":      prj.product,
                "designation":  prj.designation,
                "quantity":     prj.quantity,
                "revision":     prj.revision,
                "date":         prj.date,
                "created_by":   prj.created_by,
                "validated_by": prj.validated_by,
                "description":  prj.description,
                "divers_percent":    prj.divers_percent,
                "manual_rex_coeff":  prj.manual_rex_coeff,
            },
            "modifications": {
                "lpdc_docs": [
                    {"index": d.index, "is_selected": d.is_selected, "manual_hours": d.manual_hours}
                    for d in prj.lpdc_docs
                    if d.is_selected or d.manual_hours is not None
                ],
                "options": [
                    {"index": o.index, "is_selected": o.is_selected, "manual_hours": o.manual_hours}
                    for o in prj.options
                    if o.is_selected or o.manual_hours is not None
                ],
                "calculs": [
                    {"index": c.index, "is_selected": c.is_selected, "manual_hours": c.manual_hours}
                    for c in prj.calculs
                    if c.is_selected or c.manual_hours is not None
                ],
                "tasks": [
                    {"index": t.index, "manual_hours": t.manual_hours}
                    for t in prj.get_all_tasks()
                    if t.manual_hours is not None
                ],
                "labo": [
                    {"index": l.index, "manual_hours": l.manual_hours}
                    for l in prj.labo
                    if l.manual_hours is not None
                ],
            },
        }

    def load_project(self, data: dict):
        """Charge un projet : applique les valeurs puis les modifications."""
        prj = self.project
        pd = data.get("project", {})

        # Valeurs scalaires
        prj.crm_number   = pd.get("crm_number", "")
        prj.client       = pd.get("client", "")
        prj.affaire      = pd.get("affaire", "")
        prj.das          = pd.get("das", "")
        prj.secteur      = pd.get("secteur", "")
        prj.machine_type = pd.get("machine_type", "")
        prj.product      = pd.get("product", "")
        prj.designation  = pd.get("designation", "")
        prj.quantity     = pd.get("quantity", 1)
        prj.revision     = pd.get("revision", "A")
        prj.date         = pd.get("date", "")
        prj.created_by   = pd.get("created_by", "")
        prj.validated_by = pd.get("validated_by", "")
        prj.description  = pd.get("description", "")

        # Reconstruire les listes à partir des données sources
        prj.apply_defaults()

        # Restaurer divers/rex APRÈS apply_defaults (qui les réinitialise)
        prj.divers_percent   = pd.get("divers_percent", 0.05)
        prj.manual_rex_coeff = pd.get("manual_rex_coeff", 1.0)

        # Appliquer les modifications
        mods = data.get("modifications", {})

        lpdc_idx = {m["index"]: m for m in mods.get("lpdc_docs", [])}
        for doc in prj.lpdc_docs:
            if doc.index in lpdc_idx:
                m = lpdc_idx[doc.index]
                doc.is_selected  = m.get("is_selected", doc.is_selected)
                doc.manual_hours = m.get("manual_hours")

        opt_idx = {m["index"]: m for m in mods.get("options", [])}
        for opt in prj.options:
            if opt.index in opt_idx:
                m = opt_idx[opt.index]
                opt.is_selected  = m.get("is_selected", opt.is_selected)
                opt.manual_hours = m.get("manual_hours")

        calc_idx = {m["index"]: m for m in mods.get("calculs", [])}
        for calc in prj.calculs:
            if calc.index in calc_idx:
                m = calc_idx[calc.index]
                calc.is_selected  = m.get("is_selected", calc.is_selected)
                calc.manual_hours = m.get("manual_hours")

        task_idx = {m["index"]: m for m in mods.get("tasks", [])}
        for task in prj.get_all_tasks():
            if task.index in task_idx:
                task.manual_hours = task_idx[task.index].get("manual_hours")

        labo_idx = {m["index"]: m for m in mods.get("labo", [])}
        for labo in prj.labo:
            if labo.index in labo_idx:
                labo.manual_hours = labo_idx[labo.index].get("manual_hours")